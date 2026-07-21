from __future__ import annotations

import csv
import threading
from datetime import date
from decimal import Decimal

import pytest
from openpyxl import load_workbook

from siaf_support_toolbox.services import query_export_service
from siaf_support_toolbox.services.query_export_service import export_query_result
from siaf_support_toolbox.services.query_result_store import QueryResultStore


def test_csv_export_is_atomic_utf8_and_blocks_formula_injection(tmp_path):
    progress = []
    result = export_query_result(
        columns=("CÓDIGO", "NOME"),
        batches=(((1, "José"), (2, "=HYPERLINK('x')")),),
        output_directory=tmp_path,
        base_name="Clientes / ativos",
        file_format="csv",
        cancel_event=threading.Event(),
        on_progress=progress.append,
    )

    assert result.success
    assert result.records_processed == 2
    assert result.output_file is not None
    with result.output_file.open(encoding="utf-8-sig", newline="") as stream:
        rows = list(csv.reader(stream, delimiter=";"))
    assert rows == [
        ["CÓDIGO", "NOME"],
        ["1", "José"],
        ["2", "'=HYPERLINK('x')"],
    ]
    assert progress == [2]
    assert not tuple(tmp_path.glob("*.tmp"))


def test_csv_export_formats_dates_for_brazil(tmp_path):
    result = export_query_result(
        columns=("DATA",),
        batches=[((date(2026, 7, 19),),)],
        output_directory=tmp_path,
        base_name="Datas",
        file_format="csv",
        cancel_event=threading.Event(),
    )

    assert result.output_file is not None
    with result.output_file.open(encoding="utf-8-sig", newline="") as stream:
        rows = list(csv.reader(stream, delimiter=";"))
    assert rows == [["DATA"], ["19/07/2026"]]


def test_xlsx_export_streams_and_opens_without_recovery(tmp_path):
    result = export_query_result(
        columns=("CÓDIGO", "NOME"),
        batches=(((1, "Ana"),), ((2, "+cmd"),)),
        output_directory=tmp_path,
        base_name="Clientes",
        file_format="xlsx",
        cancel_event=threading.Event(),
    )

    assert result.success
    assert result.output_file is not None
    workbook = load_workbook(result.output_file, read_only=False, data_only=False)
    worksheet = workbook["Dados"]
    assert worksheet.freeze_panes == "A2"
    assert worksheet.auto_filter.ref == "A1:B3"
    assert tuple(worksheet.values) == (("CÓDIGO", "NOME"), (1, "Ana"), (2, "'+cmd"))
    assert worksheet["B3"].data_type == "s"
    workbook.close()


def test_csv_export_blocks_control_and_full_width_formula_prefixes(tmp_path):
    dangerous = ("\t=1+1", "\r=1+1", "\n=1+1", "\0=1+1", "＝1+1")
    result = export_query_result(
        columns=("VALOR",),
        batches=(tuple((value,) for value in dangerous),),
        output_directory=tmp_path,
        base_name="Clientes",
        file_format="csv",
        cancel_event=threading.Event(),
    )

    assert result.output_file is not None
    with result.output_file.open(encoding="utf-8-sig", newline="") as stream:
        rows = list(csv.reader(stream, delimiter=";"))
    assert tuple(row[0] for row in rows[1:]) == tuple("'" + value for value in dangerous)


def test_xlsx_export_preserves_types_from_result_cache(tmp_path):
    store = QueryResultStore(tmp_path / "cache")
    store.append_batch(((date(2026, 7, 19), Decimal("10.25"), "\0=1+1"),))

    result = export_query_result(
        columns=("DATA", "VALOR", "TEXTO"),
        batches=store.iter_batches(),
        output_directory=tmp_path / "exports",
        base_name="Produtos",
        file_format="xlsx",
        cancel_event=threading.Event(),
    )

    assert result.output_file is not None
    workbook = load_workbook(result.output_file, data_only=False)
    worksheet = workbook["Dados"]
    assert worksheet["A2"].data_type == "d"
    assert worksheet["A2"].value.date() == date(2026, 7, 19)
    assert worksheet["A2"].number_format == "DD/MM/YYYY"
    assert worksheet["B2"].value == 10.25
    assert worksheet["C2"].value == "'�=1+1"
    workbook.close()
    store.close()


def test_export_rejects_result_without_columns(tmp_path):
    with pytest.raises(ValueError, match="não possui colunas"):
        export_query_result(
            columns=(),
            batches=(),
            output_directory=tmp_path,
            base_name="Consulta cancelada",
            file_format="csv",
            cancel_event=threading.Event(),
        )

    assert not tuple(tmp_path.iterdir())


def test_canceled_export_removes_partial_file(tmp_path):
    cancel_event = threading.Event()

    def batches():
        yield ((1,),)
        cancel_event.set()
        yield ((2,),)

    result = export_query_result(
        columns=("CÓDIGO",),
        batches=batches(),
        output_directory=tmp_path,
        base_name="Produtos",
        file_format="csv",
        cancel_event=cancel_event,
    )

    assert result.canceled
    assert result.records_processed == 1
    assert not tuple(tmp_path.iterdir())


def test_xlsx_export_splits_rows_across_excel_sized_sheets(monkeypatch, tmp_path):
    monkeypatch.setattr(query_export_service, "_XLSX_MAX_ROWS", 4)
    result = export_query_result(
        columns=("CÓDIGO",),
        batches=(tuple((value,) for value in range(1, 8)),),
        output_directory=tmp_path,
        base_name="Resultado grande",
        file_format="xlsx",
        cancel_event=threading.Event(),
    )

    assert result.success
    assert result.records_processed == 7
    assert result.output_file is not None
    workbook = load_workbook(result.output_file, read_only=False)
    assert workbook.sheetnames == ["Dados", "Dados 2", "Dados 3"]
    assert tuple(workbook["Dados"].values) == (("CÓDIGO",), (1,), (2,), (3,))
    assert tuple(workbook["Dados 2"].values) == (("CÓDIGO",), (4,), (5,), (6,))
    assert tuple(workbook["Dados 3"].values) == (("CÓDIGO",), (7,))
    assert workbook["Dados"].auto_filter.ref == "A1:A4"
    assert workbook["Dados 3"].auto_filter.ref == "A1:A2"
    workbook.close()
