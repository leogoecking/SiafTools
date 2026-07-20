from __future__ import annotations

import csv
import logging
import os
import re
import tempfile
import threading
import time
import uuid
from collections.abc import Callable, Iterable
from contextlib import suppress
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

LOGGER = logging.getLogger(__name__)
_UNSAFE_FORMULA_PREFIXES = (
    "=",
    "+",
    "-",
    "@",
    "\t",
    "\r",
    "\n",
    "\0",
    "＝",
    "＋",
    "－",
    "＠",
)
_INVALID_FILENAME = re.compile(r"[^0-9A-Za-zÀ-ÿ._-]+")
_ILLEGAL_XLSX_CHARACTERS = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")


@dataclass(frozen=True, slots=True)
class QueryExportResult:
    success: bool
    file_format: str
    records_processed: int = 0
    duration_ms: int = 0
    output_file: Path | None = None
    canceled: bool = False
    error_code: str | None = None
    message: str | None = None


def export_query_result(
    *,
    columns: tuple[str, ...],
    batches: Iterable[tuple[tuple[object, ...], ...]],
    output_directory: Path,
    base_name: str,
    file_format: str,
    cancel_event: threading.Event,
    on_progress: Callable[[int], None] | None = None,
) -> QueryExportResult:
    """Exporta um resultado paginado sem materializá-lo por inteiro na memória."""
    normalized_format = file_format.casefold()
    if normalized_format not in {"csv", "xlsx"}:
        raise ValueError("Formato de exportação não suportado")
    if not columns:
        raise ValueError("O resultado não possui colunas para exportação")

    started = time.monotonic()
    output_directory.mkdir(parents=True, exist_ok=True)
    stem = _safe_stem(base_name)
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S-%f")
    destination = output_directory / (
        f"{stem}-{timestamp}-{uuid.uuid4().hex[:8]}.{normalized_format}"
    )
    descriptor, raw_temporary = tempfile.mkstemp(
        prefix=f".{stem}-", suffix=f".{normalized_format}.tmp", dir=output_directory
    )
    os.close(descriptor)
    temporary = Path(raw_temporary)
    processed = 0
    try:
        if cancel_event.is_set():
            return _canceled(normalized_format, processed, started)
        if normalized_format == "csv":
            processed = _write_csv(
                temporary, columns, batches, cancel_event, on_progress
            )
        else:
            processed = _write_xlsx(
                temporary, columns, batches, cancel_event, on_progress
            )
        if cancel_event.is_set():
            return _canceled(normalized_format, processed, started)
        os.replace(temporary, destination)
        return QueryExportResult(
            True,
            normalized_format,
            processed,
            int((time.monotonic() - started) * 1000),
            destination,
        )
    except Exception:
        LOGGER.exception("Exportação de consulta falhou")
        return QueryExportResult(
            False,
            normalized_format,
            processed,
            int((time.monotonic() - started) * 1000),
            error_code="query_export_failed",
            message="A exportação encontrou um erro interno; consulte errors.log",
        )
    finally:
        with suppress(OSError):
            temporary.unlink(missing_ok=True)


def _write_csv(
    path: Path,
    columns: tuple[str, ...],
    batches: Iterable[tuple[tuple[object, ...], ...]],
    cancel_event: threading.Event,
    on_progress: Callable[[int], None] | None,
) -> int:
    processed = 0
    with path.open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.writer(stream, delimiter=";")
        writer.writerow(columns)
        for batch in batches:
            if cancel_event.is_set():
                break
            writer.writerows(tuple(_safe_cell(value) for value in row) for row in batch)
            processed += len(batch)
            if on_progress is not None:
                on_progress(processed)
    return processed


def _write_xlsx(
    path: Path,
    columns: tuple[str, ...],
    batches: Iterable[tuple[tuple[object, ...], ...]],
    cancel_event: threading.Event,
    on_progress: Callable[[int], None] | None,
) -> int:
    workbook = Workbook(write_only=True)
    worksheet = workbook.create_sheet("Dados")
    worksheet.freeze_panes = "A2"
    header = []
    for label in columns:
        cell = WriteOnlyCell(worksheet, value=label)
        cell.font = Font(bold=True)
        header.append(cell)
    worksheet.append(header)
    for index in range(1, len(columns) + 1):
        worksheet.column_dimensions[get_column_letter(index)].width = 18

    processed = 0
    try:
        for batch in batches:
            if cancel_event.is_set():
                break
            for row in batch:
                worksheet.append([_safe_xlsx_cell(worksheet, value) for value in row])
            processed += len(batch)
            if on_progress is not None:
                on_progress(processed)
        if not cancel_event.is_set():
            final_column = get_column_letter(max(1, len(columns)))
            worksheet.auto_filter.ref = f"A1:{final_column}{processed + 1}"
            workbook.save(path)
    finally:
        workbook.close()
    return processed


def _safe_cell(value: object) -> object:
    if isinstance(value, str) and value.startswith(_UNSAFE_FORMULA_PREFIXES):
        return "'" + value
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    return value


def _safe_xlsx_cell(worksheet: object, value: object) -> object:
    if isinstance(value, datetime):
        cell = WriteOnlyCell(worksheet, value=value)
        cell.number_format = "DD/MM/YYYY HH:MM:SS"
        return cell
    if isinstance(value, date):
        cell = WriteOnlyCell(worksheet, value=value)
        cell.number_format = "DD/MM/YYYY"
        return cell
    value = _safe_cell(value)
    if isinstance(value, str):
        value = _ILLEGAL_XLSX_CHARACTERS.sub("�", value)
        if len(value) > 32767:
            return value[:32767]
    return value


def _safe_stem(value: str) -> str:
    sanitized = _INVALID_FILENAME.sub("-", value.strip()).strip(".-_")
    return (sanitized or "consulta")[:80]


def _canceled(file_format: str, processed: int, started: float) -> QueryExportResult:
    return QueryExportResult(
        False,
        file_format,
        processed,
        int((time.monotonic() - started) * 1000),
        canceled=True,
        error_code="query_export_canceled",
        message="Exportação cancelada",
    )
